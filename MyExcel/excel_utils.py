import os
import traceback
from openpyxl import load_workbook, Workbook
import pandas as pd
from OtherFunctions.logdefine import MyLogging
import xlrd
import xlwings as xw

mlogger = MyLogging(file=f"./log.log")


class MyExcel:
    def __init__(self):
        self.data = pd.DataFrame()
        self.inx2col = pd.DataFrame()
        self.col2inx = pd.DataFrame()

    def read_excels(self, input_excel):
        """
        读取excel文件
        @param input_excel: 输入excel地址
        @return:数据帧
        """
        mlogger.info(f'读取{input_excel}')
        if input_excel.endswith('.csv'):
            try:
                self.data = pd.read_csv(input_excel, encoding='gbk')
            except Exception as e:
                print(e)
                self.data = pd.read_csv(input_excel, encoding='utf-8')
        else:
            # print(pd.ExcelFile(input_excel).sheet_names)
            self.data = pd.read_excel(input_excel, sheet_name=pd.ExcelFile(input_excel).sheet_names[0])
        return self.data

    def split_excel_num(self, input_excel, split_size, output_dir):
        """
        按数拆分excel文件，存到指定目录
        :param input_excel: 输入excel地址
        :param split_size: 要拆分的数目
        :param output_dir: 输出的目录
        """
        df_input = self.read_excels(input_excel)
        num = int(df_input.shape[0] / split_size) + 1
        for i in range(num):
            mlogger.info(f'分割excel{input_excel}-{i}')
            start = i * split_size
            end = (i + 1) * split_size
            if end > df_input.shape[0]:
                end = df_input.shape[0]
            # df_input[start:end].to_excel('split_{0}.xlsx'.format(i + 1))
            df_single = df_input[start:end]
            if input_excel.endswith('.csv'):
                df_single.to_csv(f"{output_dir}/拆分后{i}-{os.path.basename(input_excel)}", encoding='gbk', index=False)
            else:
                df_single.to_excel(f"{output_dir}/拆分后{i}-{os.path.basename(input_excel)}", index=False)

    def split_excel_context(self, input_excel, split_column, output_dir):
        """
        按内容拆分excel文件，存到指定目录
        :param input_excel: 输入excel地址
        :param split_column: 要拆分的列
        :param output_dir: 输出的目录
        """
        df_input = self.read_excels(input_excel)
        unique_values = df_input[split_column].unique()
        for unique_value in unique_values:
            mlogger.info(f'分割excel{input_excel}-{unique_value}')
            df_single = df_input[df_input[split_column] == unique_value]
            if input_excel.endswith('.csv'):
                df_single.to_csv(f"{output_dir}/拆分后-{unique_value}.csv", index=False)
            else:
                df_single.to_excel(f"{output_dir}/拆分后-{unique_value}.xlsx", index=False)

    def merge_excels(self, excel_paths, output_path):
        """
        合并不同的excel到一个
        :param excel_paths: 列表，多个excel路径
        :param output_path: 合并后的excel路径
        """
        df_list = []
        for excel_path in excel_paths:
            if not os.path.exists(excel_path):
                continue
            mlogger.info(f'合并excel{excel_path}')
            df_list.append(self.read_excels(excel_path))
        df_all = pd.concat(df_list, axis=0)
        df_all.to_excel(output_path, index=False)

    @staticmethod
    def merge_xlsm(excel_paths, output_path):
        """
        合并不同的excel的不同sheet进行合并到一个
        :param excel_paths: 列表，多个excel路径
        :param output_path: 合并后的excel路径
        """
        # 获取指定目录中的所有xlsm文件路径
        # file_paths = glob.glob(os.path.join(folder_path, '*.xlsm')

        if excel_paths:
            if output_path.endswith('.xlsm'):
                wb1 = Workbook()
                wb1.save(output_path)
                wb1 = load_workbook(output_path, keep_vba=True)
                wb1.remove_sheet(wb1.active)
            else:
                wb1 = Workbook()
                wb1.remove_sheet(wb1.active)
            if len(excel_paths) > 0:
                for path in excel_paths:
                    if path.endswith('.xls'):
                        wb = xlrd.open_workbook(filename=path)
                        for sheet_name in wb.sheet_names():
                            if sheet_name not in wb1.sheetnames:
                                wb1.create_sheet(sheet_name)

                            ws1 = wb1[sheet_name]
                            ws = wb.sheet_by_name(sheet_name)

                            for row in range(ws.nrows):
                                last_row = ws1.max_row + 1
                                temp_row = []
                                # 处理每列的值
                                for col in range(ws.ncols):
                                    # print(ws1.cell(row, col).value)
                                    temp_row.append(ws.cell(row, col).value)

                                for i in range(len(temp_row)):
                                    cell = ws1.cell(row=last_row, column=i + 1)
                                    # cell = f'{chr(65 + i)}{last_row}'
                                    cell.value = temp_row[i]

                    else:
                        wb = load_workbook(filename=path, keep_vba=True)
                        # 读取每个工作表的内容
                        # sheets = [sheetname for sheetname in wb.sheetnames]
                        for sheet_name in wb.sheetnames:
                            if sheet_name not in wb1.sheetnames:
                                wb1.create_sheet(sheet_name)

                            ws1 = wb1[sheet_name]
                            ws = wb[sheet_name]

                            for row in ws.iter_rows():
                                last_row = ws1.max_row + 1
                                temp_row = []
                                # 处理每列的值
                                for cell in row:
                                    if isinstance(cell.value, (int, float)):
                                        temp_row.append(float('nan') if pd.isnull(cell.value) else cell.value)
                                    elif isinstance(cell.value, str):
                                        temp_row.append(str(cell.value))
                                    else:
                                        temp_row.append(None)
                                for i in range(len(temp_row)):
                                    cell = ws1.cell(row=last_row, column=i + 1)
                                    # cell = f'{chr(65 + i)}{last_row}'
                                    cell.value = temp_row[i]
                                    # ws[cell] = temp_row[i]

            wb1.save(output_path)

    @staticmethod
    def merge_excel_sheets(file_paths: list, header_num, save_name):
        """
        把各文件中Worksheet名相同的表格合并到一个sheet中
        :param file_paths: 给出来源工作簿的文件路径,list
        :param header_num: 表头行数
        :param save_name: 保存路径以及名称
        :return:
        """

        result_dic = {}  # {sheetName:pandas}
        for file_path in file_paths:
            app = xw.App(visible=False, add_book=False)
            workbook = app.books.open(file_path)
            sheet_names = list(reversed([j.name for j in workbook.sheets]))
            for sheet_name in sheet_names:
                mlogger.info(f'{file_path}-{sheet_name}')
                worksheet = workbook.sheets[sheet_name]
                # 以DataFrame格式读取要合并的工作表数据
                if worksheet.used_range.address.rsplit("$", 2)[1] == 'A':
                    address1 = worksheet.used_range.address.rsplit("$", 2)[0] + 'B' \
                               + worksheet.used_range.address.rsplit("$", 2)[-1]
                    value = worksheet.range(address1).options(pd.DataFrame).value  # 读取要拆分的工作表中的所有数据
                else:
                    value = worksheet.range(worksheet.used_range).options(pd.DataFrame).value  # 读取要拆分的工作表中的所有数据
                value_header = value.iloc[0:header_num - 1, :]  # 获取表头
                value_data = value.iloc[header_num - 1:value.shape[0], :]  # 获取表格内容

                # 如果工作簿的工作表名不存在，则新增，否则合并替换
                if sheet_name not in result_dic.keys():
                    result_dic[sheet_name] = pd.concat([value_header, value_data], join='outer', axis=0)
                else:
                    result_dic[sheet_name] = pd.concat([result_dic[sheet_name], value_data], join='outer', axis=0)
            workbook.close()
            app.quit()

        if result_dic:
            new = xw.App(visible=False, add_book=False)
            new_wb = new.books.add()
            for sheetName, value1 in result_dic.items():
                mlogger.info(f'正在合并{sheet_name}')
                new_worksheet = new_wb.sheets.add(sheetName)  # 在工作簿中新增工作表并命名为当前的产品名称
                new_worksheet.range('A1').options(index=True).value = value1  # 将按分组好的数据添加到新增的工作表
            Activate_worksheet = new_wb.sheets['Sheet1']
            Activate_worksheet.delete()
            new_wb.save(save_name)
            new_wb.close()
            new.quit()

    @staticmethod
    def merge_excel_sheet(file_paths: list, header_num, save_name):
        """
        把各文件中Worksheet名相同的表格合并到一个sheet中
        :param file_paths: 给出来源工作簿的文件路径,list
        :param header_num: 表头行数
        :param save_name: 保存路径以及名称
        :return:
        """

        result_dic = {}
        for file_path in file_paths:
            app = xw.App(visible=False, add_book=False)
            workbook = app.books.open(file_path)
            sheet_names = [j.name for j in workbook.sheets]
            for sheet_name in sheet_names:
                mlogger.info(f'{file_path}-{sheet_name}')
                worksheet = workbook.sheets[sheet_name]
                # 以DataFrame格式读取要合并的工作表数据
                if worksheet.used_range.address.rsplit("$", 2)[1] == 'A':
                    address1 = worksheet.used_range.address.rsplit("$", 2)[0] + 'B' \
                               + worksheet.used_range.address.rsplit("$", 2)[-1]
                    value = worksheet.range(address1).options(pd.DataFrame).value  # 读取要拆分的工作表中的所有数据
                else:
                    value = worksheet.range(worksheet.used_range).options(pd.DataFrame).value  # 读取要拆分的工作表中的所有数据
                value_header = value.iloc[0:header_num - 1, :]  # 获取表头
                value_data = value.iloc[header_num - 1:value.shape[0], :]  # 获取表格内容

                # 如果工作簿的工作表名不存在，则新增，否则合并替换
                if 'merge' not in result_dic.keys():
                    result_dic['merge'] = pd.concat([value_header, value_data], join='outer', axis=0)
                else:
                    print(result_dic['merge'])
                    print(value_data)
                    result_dic['merge'] = pd.concat([result_dic['merge'], value_data], join='outer', axis=0)
            workbook.close()
            app.quit()

        if result_dic:
            new = xw.App(visible=False, add_book=False)
            new_wb = new.books.add()
            for sheetName, value1 in result_dic.items():
                mlogger.info(f'正在合并{sheet_name}')
                new_worksheet = new_wb.sheets.add(sheetName)  # 在工作簿中新增工作表并命名为当前的产品名称
                new_worksheet.range('A1').options(index=True).value = value1  # 将按分组好的数据添加到新增的工作表

            Activate_worksheet = new_wb.sheets['Sheet1']
            Activate_worksheet.delete()
            new_wb.save(save_name)
            new_wb.close()
            new.quit()

    @staticmethod
    def split_excel(file_path, header_num, split_name, save_path):
        """
        按条件将一个工作表拆分为多个工作表
        :param file_path: 给出来源工作簿的文件路径及工作簿名称
        :param header_num: 表头行数
        :param split_name: 拆分列名称
        :param save_path: 保存路径
        :return:
        """
        app = xw.App(visible=False, add_book=False)
        workbook = app.books.open(file_path)
        sheet_names = list(reversed([j.name for j in workbook.sheets]))

        result_dic = {}  # {index:{sheetName:pandas}}
        result_dic_NON = {}  # {sheetName:pandas}
        for sheet_name in sheet_names:
            worksheet = workbook.sheets[sheet_name]
            # 以DataFrame格式读取要拆分的工作表数据
            value = worksheet.range(worksheet.used_range).options(pd.DataFrame).value  # 读取要拆分的工作表中的所有数据
            value_header = value.iloc[0:header_num - 1, :]  # 获取表头
            value_data = value.iloc[header_num - 1:value.shape[0], :]  # 获取表格内容
            # 将数据按照“名称”分组
            if split_name in value_data.columns:
                data = value_data.groupby(split_name)
                for idx, group in data:
                    # 如果工作簿的工作表名不存在，则新增，否则替换
                    if idx not in result_dic.keys():
                        result_dic[idx] = {}
                        result_dic[idx][sheet_name] = pd.concat([value_header, group], join='outer', axis=0)
                    else:
                        if sheet_name not in result_dic[idx].keys():
                            result_dic[idx][sheet_name] = pd.concat([value_header, group], join='outer', axis=0)
                        else:
                            result_dic[idx][sheet_name] = pd.concat([result_dic[idx][sheet_name], group], join='outer',
                                                                    axis=0)

            else:
                if sheet_name not in result_dic_NON.keys():
                    result_dic_NON[sheet_name] = pd.concat([value_header, value_data], join='outer', axis=0)
                else:
                    result_dic_NON[sheet_name] = pd.concat([result_dic_NON[sheet_name], value_data], join='outer',
                                                           axis=0)

        workbook.close()
        app.quit()

        if result_dic:
            for idx, sheetName in result_dic.items():
                new = xw.App(visible=False, add_book=False)
                new_wb = new.books.add()

                if sheetName:
                    for sheet1, value1 in sheetName.items():
                        new_worksheet = new_wb.sheets.add(sheet1)  # 在工作簿中新增工作表并命名为当前的产品名称
                        new_worksheet.range('A1').options(index=False).value = value1  # 将按分组好的数据添加到新增的工作表
                if result_dic_NON:
                    for sheet_NON, value_NON in result_dic_NON.items():
                        new_worksheet = new_wb.sheets.add(sheet_NON)  # 在工作簿中新增工作表并命名为当前的产品名称
                        new_worksheet.range('A1').options(index=False).value = value_NON  # 将按分组好的数据添加到新增的工作表

                Activate_worksheet = new_wb.sheets['Sheet1']
                Activate_worksheet.delete()
                save_name = f"{os.path.basename(file_path).rsplit('.', 1)[0]}-" \
                            f"{idx}.{os.path.basename(file_path).rsplit('.', 1)[-1]}"

                new_wb.save(os.path.join(save_path, save_name))
                print(os.path.join(save_path, save_name))
                new_wb.close()
                new.quit()

    @staticmethod
    def split_excel_number(file_path, header_num, split_num, save_path):
        """
        按条件将一个工作表拆分为多个工作表
        :param file_path: 给出来源工作簿的文件路径及工作簿名称
        :param header_num: 表头行数
        :param split_num: 拆分行数
        :param save_path: 保存路径
        :return:
        """
        app = xw.App(visible=False, add_book=False)
        workbook = app.books.open(file_path)
        sheet_names = list(reversed([j.name for j in workbook.sheets]))

        result_dic = {}  # {index:{sheetName:pandas}}
        for sheet_name in sheet_names:
            worksheet = workbook.sheets[sheet_name]
            # 以DataFrame格式读取要拆分的工作表数据
            value = worksheet.range(worksheet.used_range).options(pd.DataFrame).value  # 读取要拆分的工作表中的所有数据
            value_header = value.iloc[0:header_num - 1, :]  # 获取表头
            value_data = value.iloc[header_num - 1:value.shape[0], :]  # 获取表格内容
            # 将数据行数拆分
            n = 0
            while n*split_num < value_data.shape[0]:
                s = n*split_num
                e = (n+1) * split_num
                if e > value_data.shape[0]:
                    e = value_data.shape[0]

                data = value_data.iloc[s:e, :]
                if n not in result_dic.keys():
                    result_dic[n] = {}
                    result_dic[n][sheet_name] = pd.concat([value_header, data], join='outer', axis=0)
                else:
                    if sheet_name not in result_dic[n].keys():
                        result_dic[n][sheet_name] = pd.concat([value_header, data], join='outer', axis=0)
                    else:
                        result_dic[n][sheet_name] = pd.concat([result_dic[n][sheet_name], value_header, data], join='outer', axis=0)
                n += 1
        workbook.close()
        app.quit()

        if result_dic:
            for idx, sheetName in result_dic.items():
                new = xw.App(visible=False, add_book=False)
                new_wb = new.books.add()

                if sheetName:
                    for sheet1, value1 in sheetName.items():
                        new_worksheet = new_wb.sheets.add(sheet1)  # 在工作簿中新增工作表并命名为当前的产品名称
                        new_worksheet.range('A1').options(index=False).value = value1  # 将按分组好的数据添加到新增的工作表

                Activate_worksheet = new_wb.sheets['Sheet1']
                Activate_worksheet.delete()
                save_name = f"{os.path.basename(file_path).rsplit('.', 1)[0]}-" \
                            f"{idx}.{os.path.basename(file_path).rsplit('.', 1)[-1]}"

                new_wb.save(os.path.join(save_path, save_name))
                new_wb.close()
                new.quit()

    def split_excels(self, excel_paths, header_num, split_name, save_path, on):
        if on == 'text':
            for i in excel_paths:
                self.split_excel(i, header_num, split_name, save_path)
        else:
            for i in excel_paths:
                self.split_excel_number(i, header_num, split_name, save_path)

    def do_col2inx(self, input_excel, output_dir, id_var: list, value_var: list):
        """
        列转行
        @param input_excel:输入excel地址
        @param output_dir:输出路径
        @param id_var:指定哪些列应该保留在结果中,而不是被转换为变量。默认情况下,所有未指定的列都将被转换为变量。可以使用列名或列索引指定。
        @param value_var:指定哪些列应该被转换为变量。可以使用列名或列索引指定。如果未指定,则使用除 id_vars 列以外的所有列作为变量。
        @return:
        """
        if not value_var or not id_var:
            return None

        if not id_var or not value_var:
            return
        mlogger.info(f'excel列转行')
        self.read_excels(input_excel)
        try:
            self.col2inx = pd.melt(self.data, id_vars=id_var, value_vars=value_var).dropna()
            for i in range(int(self.col2inx.shape[0]/1048576 + 1)):
                start = i*1048576
                end = start + 1048575
                if end > self.col2inx.shape[0]:
                    end = self.col2inx.shape[0]
                res = self.col2inx.iloc[start:end]
                res.to_excel(
                    f"{output_dir}/列转行-{os.path.basename(input_excel).rsplit('.', 1)[0]}"
                    f"-{i}.{os.path.basename(input_excel).rsplit('.', 1)[-1]}",
                    index=False)
            mlogger.info(f'excel列转行完成')
        except Exception as e:
            print(e)
            mlogger.info(f'excel列转行失败\n{traceback.format_exc()}')

    def do_inx2col(self, input_excel, output_dir, id_var: list, value_var: str):
        """
        行转列
        @param input_excel: 输入excel地址
        @param output_dir: 输出路径
        @param id_var: 指定哪些列（可多列）应该保留在结果中,而不是被转换为变量。
        @param value_var: 指定某列（单列）应该被转换为变量。
        @return:
        """
        if not id_var:
            return
        mlogger.info(f'excel行转列')
        self.read_excels(input_excel)
        try:
            self.data['NAME_NBR'] = self.data.groupby(id_var).cumcount() + 1  # 计算第几次重复出现
            self.inx2col = self.data.pivot(index=id_var, columns='NAME_NBR', values=value_var)
            self.inx2col.columns = [f'value{c}' for c in self.inx2col.columns]
            self.inx2col.reset_index(inplace=True)
            self.inx2col.to_excel(f"{output_dir}/行转列-{os.path.basename(input_excel)}", index=False)
            mlogger.info(f'excel行转列完成')
        except Exception as e:
            print(e)
            mlogger.info(f'excel行转列失败\n{traceback.format_exc()}')


if __name__ == "__main__":
    # excel_paths = [r"D:\workbench\ant-python-pyqt6\chapter08_splitexcel\output\拆分后-单肩包.xlsx",
    #                r"D:\workbench\ant-python-pyqt6\chapter08_splitexcel\output\拆分后-背包.xlsx"]
    # output_path = "合并后的excel.xlsx"
    # merge_excels(excel_paths, output_path)
    qqq = [os.path.join(r'F:\工作\cell', i) for i in os.listdir(r'F:\工作\cell')]

    q = MyExcel()
    q.merge_excels(qqq, r'F:\工作\cell\xx.xlsx')
    # q.do_inx2col( input_excel, output_dir, id_var: list, value_var: str)
    q.do_inx2col(r"E:\Desktop\和.xlsx", r"E:\Desktop", ['任务负责人'], '重要工作事项-分类')

    print(q.inx2col)
