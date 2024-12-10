import os
import sqlite3
import pandas as pd
import traceback
from openpyxl import load_workbook
import warnings
import time
from OtherFunctions.logdefine import MyLogging

warnings.filterwarnings('ignore')
t = time.strftime("%Y%m%d")
mlogger = MyLogging(file=f"./log.log")


def import_cell(path):
    mlogger.info(f'导入小区信息')
    try:
        conn = sqlite3.connect('add_nei.db')
        sheetnames = pd.ExcelFile(path).sheet_names
        if '4G' in sheetnames:
            cell = pd.read_excel(path, sheet_name='4G')
            cell.to_sql('lte', conn, if_exists='replace', index=False)

        if '5G' in sheetnames:
            cell = pd.read_excel(path, sheet_name='5G')
            cell.to_sql('nr', conn, if_exists='replace', index=False)
        conn.close()
    except:
        mlogger.info(f'{traceback.format_exc()}')


def import_nei(path):
    mlogger.info(f'导入邻区对')
    try:
        conn = sqlite3.connect('add_nei.db')
        cell = pd.read_excel(path, sheet_name='4G-4G')
        cell.to_sql('lte2lte', conn, if_exists='replace', index=False)
        cell = pd.read_excel(path, sheet_name='5G-5G')
        cell.to_sql('nr2nr', conn, if_exists='replace', index=False)
        cell = pd.read_excel(path, sheet_name='5G-4G')
        cell.to_sql('nr2lte', conn, if_exists='replace', index=False)
        conn.close()
    except:
        mlogger.info(f'{traceback.format_exc()}')


def do_add_nei(output):
    conn = sqlite3.connect('add_nei.db')
    cursor = conn.cursor()
    for i in ['lte2lte', 'nr2nr', 'nr2lte']:
        cursor.execute(f"SELECT EXISTS(SELECT 1 FROM '{i}' LIMIT 1)")
        result = cursor.fetchone()
        # 检查结果
        if result[0]:
            print(f"表 {i} 有数据。")
            if i == 'lte2lte':
                t0 = i.split('2')[0]
                str1 = f"""SELECT {t0}.网管, {t0}.子网ID,{t0}.网元ID,{t0}.ENODEBID,{t0}.小区标识,{t0}.站名,{i}.N_enb,{i}.N_cell,
    N.小区名,N.PCI,N.TAC,N.频段指示,N.上行中心频率,N.下行中心频率,N.带宽,N.PLMN列表,{t0}.PLMN列表 as S_PLMN列表 FROM {i} 
    LEFT JOIN {t0} ON {i}.S_enb = {t0}.ENODEBID and {i}.S_cell = {t0}.小区标识
    LEFT JOIN {t0} as N ON {i}.N_enb = N.ENODEBID and {i}.N_cell = N.小区标识;"""

                df1 = pd.read_sql_query(str1, conn)
                df_na = df1[df1['PLMN列表'].isna() | df1['S_PLMN列表'].isna()]
                df_d = df1[~(df1['PLMN列表'].isna() | df1['S_PLMN列表'].isna())]
                df_p = pd.read_excel(os.path.abspath(r'.\config\MyNeighbor\PLMN配置.xlsx'))
                df_d = df_d.merge(df_p, left_on=['PLMN列表', 'S_PLMN列表'], right_on=['N', 'S'])
                df_na.to_excel(os.path.join(output, 'lte2lte_小区信息缺失.xlsx'))

                for v in list(set(df_d['网管'].values)):
                    df_v = df_d[df_d['网管'] == v]
                    if v.__contains__('U31'):
                        filename = os.path.abspath(r'.\config\MyNeighbor\电联共享\Excel_addcellrelation_lte.xlsm')
                        wb = load_workbook(filename=filename, keep_vba=True)
                        df_ex = pd.DataFrame(
                            columns=["Result", "MODIND", "SubNetwork", "MEID", "ExternalEUtranCellFDD", "srcENBId",
                                     "mcc", "mnc", "eNBId", "cellLocalId", "plmnIdList", "userLabel", "freqBandInd",
                                     "earfcnUl", "earfcnDl", "pci", "tac", "bandWidthDl", "bandWidthUl"])
                        df_ex['SubNetwork'] = df_v['子网ID']
                        df_ex['Result'] = ''
                        df_ex['ExternalEUtranCellFDD'] = ''
                        df_ex['MODIND'] = 'A'
                        df_ex['MEID'] = df_v['网元ID']
                        df_ex['srcENBId'] = df_v['ENODEBID']
                        df_ex['mcc'] = df_v['P'].apply(lambda x: x.split(';')[0].split(',')[0])
                        df_ex['mnc'] = df_v['P'].apply(lambda x: x.split(';')[0].split(',')[-1])
                        df_ex['eNBId'] = df_v['N_enb']
                        df_ex['cellLocalId'] = df_v['N_cell']
                        df_ex['plmnIdList'] = df_v['P']
                        df_ex['userLabel'] = df_v['小区名']
                        df_ex['freqBandInd'] = df_v['频段指示'].apply(lambda x: int(x))
                        df_ex['earfcnUl'] = df_v['上行中心频率']
                        df_ex['earfcnDl'] = df_v['下行中心频率']
                        df_ex['pci'] = df_v['PCI'].apply(lambda x: int(x))
                        df_ex['tac'] = df_v['TAC'].apply(lambda x: int(x))
                        df_ex['bandWidthDl'] = df_v['带宽'].apply(lambda x: int(x))
                        df_ex['bandWidthUl'] = df_v['带宽'].apply(lambda x: int(x))
                        df_ex.drop_duplicates(inplace=True)
                        df_ex = df_ex[df_ex['srcENBId'] != df_ex['eNBId']]

                        sheet = wb['ExternalEUtranCellFDD']
                        r = 6
                        for va in [j.tolist() for j in df_ex.values]:
                            c = 1
                            for vv in va:
                                sheet.cell(r, c).value = vv
                                c += 1
                            r += 1
                        df_nei = pd.DataFrame(
                            columns=["Result", "MODIND", "SubNetwork", "MEID", "EUtranRelation", 'CellId', "srcENBId",
                                     "mcc", "mnc", "eNBId", "NCellId", "userLabel"])
                        df_nei['SubNetwork'] = df_v['子网ID']
                        df_nei['Result'] = ''
                        df_nei['EUtranRelation'] = ''
                        df_nei['MODIND'] = 'A'
                        df_nei['MEID'] = df_v['网元ID']
                        df_nei['CellId'] = df_v['小区标识']
                        df_nei['srcENBId'] = df_v['ENODEBID']
                        df_nei['mcc'] = df_v['P'].apply(lambda x: x.split(';')[0].split(',')[0])
                        df_nei['mnc'] = df_v['P'].apply(lambda x: x.split(';')[0].split(',')[-1])
                        df_nei['eNBId'] = df_v['N_enb']
                        df_nei['NCellId'] = df_v['N_cell']
                        df_nei['userLabel'] = df_v['小区名']
                        df_nei.drop_duplicates(inplace=True)
                        sheet = wb['EUtranRelation']
                        r = 6
                        for va in [j.tolist() for j in df_nei.values]:
                            c = 1
                            for vv in va:
                                sheet.cell(r, c).value = vv
                                c += 1
                            r += 1
                        wb.save(os.path.join(output, 'Excel_addcellrelation_lte.xlsm'))
                    if v.__contains__('UME'):
                        filename = os.path.abspath(r'.\config\MyNeighbor\电联共享\RANCM-addcellrelation_lte.xlsx')
                        wb = load_workbook(filename=filename)
                        df_ex = pd.DataFrame(
                            columns=["MODIND", "ManagedElementType", "SubNetwork", "ManagedElement", "NE_Name",
                                     "ENBCUCPFunction_moId", "mcc", "mnc", "eNBId", "CellId", "userLabel", "cellType",
                                     "freqBandInd", "earfcnUl", "earfcnDl", "pci", "tac", "mcclist", "mnclist",
                                     "bandWidthDl", "bandWidthUl"])
                        df_ex['SubNetwork'] = df_v['子网ID']
                        df_ex['MODIND'] = 'A'
                        df_ex['ManagedElementType'] = 'ITBBU'
                        df_ex['ManagedElement'] = df_v['网元ID']
                        df_ex['NE_Name'] = ''
                        df_ex['ENBCUCPFunction_moId'] = df_v['ENODEBID']
                        df_ex['mcc'] = df_v['P'].apply(lambda x: x.split(';')[0].split(',')[0])
                        df_ex['mnc'] = df_v['P'].apply(lambda x: x.split(';')[0].split(',')[-1])
                        df_ex['eNBId'] = df_v['N_enb']
                        df_ex['CellId'] = df_v['N_cell']
                        df_ex['userLabel'] = df_v['小区名']
                        df_ex['cellType'] = '0'
                        df_ex['freqBandInd'] = df_v['频段指示'].apply(lambda x: int(x))
                        df_ex['earfcnUl'] = df_v['上行中心频率']
                        df_ex['earfcnDl'] = df_v['下行中心频率']
                        df_ex['pci'] = df_v['PCI'].apply(lambda x: int(x))
                        df_ex['tac'] = df_v['TAC'].apply(lambda x: int(x))
                        df_ex['mcclist'] = df_v['P'].apply(lambda x: str(x.split(';')[0].split(',')[0]) + ';' + str(
                            x.split(';')[1].split(',')[0]) if len(x.split(';')) == 2 else str(x.split(',')[0]))
                        df_ex['mnclist'] = df_v['P'].apply(lambda x: str(x.split(';')[0].split(',')[-1]) + ';' + str(
                            x.split(';')[1].split(',')[-1]) if len(x.split(';')) == 2 else str(x.split(',')[-1]))
                        df_ex['bandWidthDl'] = df_v['带宽'].apply(lambda x: int(x))
                        df_ex['bandWidthUl'] = df_v['带宽'].apply(lambda x: int(x))
                        df_ex.drop_duplicates(inplace=True)
                        df_ex = df_ex[df_ex['ENBCUCPFunction_moId'] != df_ex['eNBId']]
                        sheet = wb['ExternalEUtranCellFDDLTE']
                        r = 6
                        for va in [j.tolist() for j in df_ex.values]:

                            c = 1
                            for vv in va:
                                sheet.cell(r, c).value = vv
                                c += 1
                            r += 1

                        df_nei = pd.DataFrame(
                            columns=["MODIND", "ManagedElementType", "SubNetwork", "ManagedElement", "NE_Name",
                                     'sourceEnbId', "CellId",
                                     "mcc", "mnc", "targetEnbId", "targetCellId", 'additionMode', 'userLabel'])

                        df_nei['SubNetwork'] = df_v['子网ID']
                        df_nei['ManagedElement'] = df_v['网元ID']
                        df_nei['MODIND'] = 'A'
                        df_nei['ManagedElementType'] = 'ITBBU'
                        df_nei['NE_Name'] = ''
                        df_nei['sourceEnbId'] = df_v['ENODEBID']
                        df_nei['CellId'] = df_v['小区标识']
                        df_nei['mcc'] = df_v['P'].apply(lambda x: x.split(';')[0].split(',')[0])
                        df_nei['mnc'] = df_v['P'].apply(lambda x: x.split(';')[0].split(',')[-1])
                        df_nei['targetEnbId'] = df_v['N_enb']
                        df_nei['targetCellId'] = df_v['N_cell']
                        df_nei['userLabel'] = df_v['小区名']
                        df_nei['additionMode'] = '单配[0]'
                        df_nei.drop_duplicates(inplace=True)
                        sheet = wb['EUtranRelationFDDLTE']
                        r = 6
                        for va in [j.tolist() for j in df_nei.values]:
                            c = 1
                            for vv in va:
                                sheet.cell(r, c).value = vv
                                c += 1
                            r += 1
                        wb.save(os.path.join(output, 'RANCM-addcellrelation_lte.xlsx'))
            elif i == 'nr2lte':
                t0 = i.split('2')[0]
                t1 = i.split('2')[-1]
                str1 = f"""SELECT {t0}.子网ID,{t0}.网元ID,{t0}.GNBID,{t0}.小区标识,{t0}.网元名称,{i}.N_enb,{i}.N_cell,
    N.小区名,N.PCI,N.TAC,N.频段指示,N.上行中心频率,N.下行中心频率,N.带宽,N.PLMN列表,{t0}.PLMN列表 as S_PLMN列表 FROM {i} 
    LEFT JOIN {t0} ON {i}.S_enb = {t0}.GNBID and {i}.S_cell = {t0}.小区标识
    LEFT JOIN {t1} as N ON {i}.N_enb = N.ENODEBID and {i}.N_cell = N.小区标识;"""

                df1 = pd.read_sql_query(str1, conn)
                df_na = df1[df1['PLMN列表'].isna() | df1['S_PLMN列表'].isna()]
                df_d = df1[~(df1['PLMN列表'].isna() | df1['S_PLMN列表'].isna())]
                df_p = pd.read_excel(os.path.abspath(r'.\config\MyNeighbor\PLMN配置.xlsx'))
                df_v = df_d.merge(df_p, left_on=['PLMN列表', 'S_PLMN列表'], right_on=['N', 'S'])
                df_na.to_excel(os.path.join(output, 'lte2lte_小区信息缺失.xlsx'))

                filename = os.path.abspath(r'.\config\MyNeighbor\电联共享\RANCM-addcellrelation_nr.xlsx')
                wb = load_workbook(filename=filename)
                df_ex = pd.DataFrame(
                    columns=["MODIND", "ManagedElementType", "SubNetwork", "ManagedElement", "NE_Name",
                             "sourcePLMNId", "sourceGNBId", "eNBId", "pLMNId", "cellLocalId", "userLabel", "pLMNIdList",
                             "pci", "tac", "bandIndicator", "frequencyUL_or_ulEarfcn", "frequencyDL_or_dlEarfcn",
                             "ulBandWidth", "dlBandWidth"])
                df_ex['SubNetwork'] = df_v['子网ID']
                df_ex['MODIND'] = 'A'
                df_ex['ManagedElementType'] = 'ITBBU'
                df_ex['ManagedElement'] = df_v['网元ID']
                df_ex['NE_Name'] = ''
                df_ex['sourcePLMNId'] = df_v['S_PLMN列表'].apply(lambda x: x.split(';')[0])
                df_ex['sourceGNBId'] = df_v['GNBID']
                df_ex['eNBId'] = df_v['N_enb']
                df_ex['pLMNId'] = df_v['P'].apply(lambda x: x.split(';')[0])

                df_ex['cellLocalId'] = df_v['N_cell']
                df_ex['userLabel'] = df_v['小区名']
                df_ex['pLMNIdList'] = df_v['P']
                df_ex['bandIndicator'] = df_v['频段指示'].apply(lambda x: int(x))
                df_ex['frequencyUL_or_ulEarfcn'] = df_v['上行中心频率']
                df_ex['frequencyDL_or_dlEarfcn'] = df_v['下行中心频率']
                df_ex['pci'] = df_v['PCI'].apply(lambda x: int(x))
                df_ex['tac'] = df_v['TAC'].apply(lambda x: int(x))
                df_ex['ulBandWidth'] = df_v['带宽'].apply(lambda x: int(x))
                df_ex['dlBandWidth'] = df_v['带宽'].apply(lambda x: int(x))
                df_ex.drop_duplicates(inplace=True)
                sheet = wb['ExternalEutranCellFDD']
                r = 6
                for va in [j.tolist() for j in df_ex.values]:

                    c = 1
                    for vv in va:
                        sheet.cell(r, c).value = vv
                        c += 1
                    r += 1

                df_nei = pd.DataFrame(
                    columns=["MODIND", "ManagedElementType", "sourcePlmn", "sourceGNBId", "sourceCellId",
                             'targetPlmn', "targetENBId", 'targetCellId', 'userLabel', 'targetCellType'])

                df_nei['sourcePlmn'] = df_v['S_PLMN列表'].apply(lambda x: x.split(';')[0])
                df_nei['sourceGNBId'] = df_v['GNBID']
                df_nei['MODIND'] = 'A'
                df_nei['ManagedElementType'] = 'ITBBU'
                df_nei['sourceCellId'] = df_v['小区标识']
                df_nei['targetPlmn'] = df_v['PLMN列表'].apply(lambda x: x.split(';')[0])
                df_nei['targetENBId'] = df_v['N_enb']
                df_nei['targetCellId'] = df_v['N_cell']
                df_nei['userLabel'] = df_v['小区名']
                df_nei['targetCellType'] = df_v.apply(lambda x: 'TDD' if x.上行中心频率 == x.下行中心频率 else 'FDD', axis=1)
                df_nei.drop_duplicates(inplace=True)
                sheet = wb['EutranCellRelation']
                r = 6
                for va in [j.tolist() for j in df_nei.values]:
                    c = 1
                    for vv in va:
                        sheet.cell(r, c).value = vv
                        c += 1
                    r += 1
                wb.save(os.path.join(output, 'RANCM-addcellrelation_nr2lte.xlsx'))
            elif i == 'nr2nr':
                t0 = i.split('2')[0]
                t1 = i.split('2')[-1]
                str1 = f"""SELECT {t0}.子网ID,{t0}.网元ID,{t0}.GNBID,{t0}.小区标识,{t0}.网元名称,{i}.N_enb,{i}.N_cell,
    N.用户标识,N.PCI,N.TAC,N.frequencyBandList,N.双工方式,N.系统带宽,N.NRFreq的SSB的频点,N.上行载波的中心频点,
    N.下行载波的中心频点,N.上行PointA频点,N.下行PointA频点,N.子载波间隔,N.时域图谱位置,N.SSB测量位图,
    N.PLMN列表,{t0}.PLMN列表 as S_PLMN列表 FROM {i} 
    LEFT JOIN {t0} ON {i}.S_enb = {t0}.GNBID and {i}.S_cell = {t0}.小区标识
    LEFT JOIN {t1} as N ON {i}.N_enb = N.GNBID and {i}.N_cell = N.小区标识;"""

                df1 = pd.read_sql_query(str1, conn)
                df_na = df1[df1['PLMN列表'].isna() | df1['S_PLMN列表'].isna()]
                df_d = df1[~(df1['PLMN列表'].isna() | df1['S_PLMN列表'].isna())]
                df_p = pd.read_excel(os.path.abspath(r'.\config\MyNeighbor\PLMN配置.xlsx'))
                df_v = df_d.merge(df_p, left_on=['PLMN列表', 'S_PLMN列表'], right_on=['N', 'S'])
                df_na.to_excel(os.path.join(output, 'nr2nr_小区信息缺失.xlsx'))

                filename = os.path.abspath(r'.\config\MyNeighbor\电联共享\RANCM-addcellrelation_nr.xlsx')
                wb = load_workbook(filename=filename)
                df_ex = pd.DataFrame(
                    columns=["MODIND", "ManagedElementType", "SubNetwork", "ManagedElement", "NE_Name", "sourcePLMNId",
                             "sourceGNBId", "targetPLMNId", "targetGNBId", "gNBIdLength", "targetCellLocalId",
                             "userLabel", "pLMNIdList", "coverageType", "ranac", "duplexMode", "frequencyUL",
                             "bandwidthUL", "subcarrierSpacingUL", "offsetToCarrierUL", "freqBandListUL", "frequencyDL",
                             "bandwidthDL", "freqBandListDL", "subcarrierSpacingDL", "offsetToCarrierDL",
                             "pointAFrequencyDL", "pointAFrequencyUL", "refNRFreq",
                             "FrequencyBandList_freqBandIndicator", "nRPCI", "tac", "configuredEpsTAC",
                             "supportNetworkType"])
                df_ex['SubNetwork'] = df_v['子网ID']
                df_ex['MODIND'] = 'A'
                df_ex['ManagedElementType'] = 'ITBBU'
                df_ex['ManagedElement'] = df_v['网元ID']
                df_ex['NE_Name'] = df_v['网元名称']
                df_ex['sourcePLMNId'] = df_v['S_PLMN列表'].apply(lambda x: x.split(';')[0].replace(',', '-'))
                df_ex['sourceGNBId'] = df_v['GNBID']
                df_ex['targetPLMNId'] = df_v['PLMN列表'].apply(lambda x: x.split(';')[0].replace(',', '-'))
                df_ex['targetGNBId'] = df_v['N_enb']
                df_ex['gNBIdLength'] = '24'

                df_ex['targetCellLocalId'] = df_v['N_cell']
                df_ex['userLabel'] = df_v['用户标识']
                df_ex['pLMNIdList'] = df_v['P']
                df_ex['coverageType'] = '宏小区[Macro]'
                df_ex['ranac'] = '0'
                df_ex['duplexMode'] = df_v['双工方式'].apply(lambda x: '[' + x + ']')
                df_ex['frequencyUL'] = df_v['上行载波的中心频点']
                df_ex['bandwidthUL'] = df_v['系统带宽']
                df_ex['subcarrierSpacingUL'] = df_v['子载波间隔'].apply(lambda x: '[' + x + ']')
                df_ex['offsetToCarrierUL'] = '0'
                df_ex['freqBandListUL'] = df_v['frequencyBandList']
                df_ex['frequencyDL'] = df_v['下行载波的中心频点']
                df_ex['bandwidthDL'] = df_v['系统带宽']
                df_ex['subcarrierSpacingDL'] = df_v['子载波间隔'].apply(lambda x: '[' + x + ']')
                df_ex['offsetToCarrierDL'] = '0'
                df_ex['freqBandListDL'] = df_v['frequencyBandList']

                df_ex['pointAFrequencyDL'] = df_v['下行PointA频点']
                df_ex['pointAFrequencyUL'] = df_v['上行PointA频点']
                df_ex['refNRFreq'] = df_v['NRFreq的SSB的频点']
                df_ex['nRPCI'] = df_v['PCI'].apply(lambda x: int(x))
                df_ex['tac'] = df_v['TAC'].apply(lambda x: int(x))
                df_ex['configuredEpsTAC'] = '0'
                df_ex['supportNetworkType'] = 'SA[0]'
                df_ex.drop_duplicates(inplace=True)
                df_ex = df_ex[df_ex['sourceGNBId'] != df_ex['targetGNBId']]
                sheet = wb['ExternalNRCellCU']
                r = 6
                for va in [j.tolist() for j in df_ex.values]:

                    c = 1
                    for vv in va:
                        sheet.cell(r, c).value = vv
                        c += 1
                    r += 1

                df_ssb = pd.DataFrame(
                    columns=["MODIND", "ManagedElementType", "SubNetwork", "ManagedElement", "NE_Name",
                             'sourcePLMNId', "sourceGNBId", 'targetPLMNId', 'targetGNBId', 'targetCellLocalId', 'moId',
                             'smtc1PrdAndOffset', 'sf', 'ssbToMeasure', 'bitmap', 'frequency', 'subcarrierSpacing'])

                df_ssb['SubNetwork'] = df_v['子网ID']
                df_ssb['MODIND'] = 'A'
                df_ssb['ManagedElementType'] = 'ITBBU'
                df_ssb['ManagedElement'] = df_v['网元ID']
                df_ssb['NE_Name'] = df_v['网元名称']
                df_ssb['sourcePLMNId'] = df_v['S_PLMN列表'].apply(lambda x: x.split(';')[0].replace(',', '-'))
                df_ssb['sourceGNBId'] = df_v['GNBID']
                df_ssb['targetPLMNId'] = df_v['PLMN列表'].apply(lambda x: x.split(';')[0].replace(',', '-'))
                df_ssb['targetGNBId'] = df_v['N_enb'].apply(lambda x: str(x))
                df_ssb['targetCellLocalId'] = df_v['N_cell'].apply(lambda x: str(x))
                df_ssb['moId'] = df_ssb['targetGNBId'] + df_ssb['targetCellLocalId']
                df_ssb['smtc1PrdAndOffset'] = 'SMTC周期20ms[sf20]'
                df_ssb['sf'] = '0'
                df_ssb['ssbToMeasure'] = df_v['SSB测量位图']
                df_ssb['bitmap'] = df_v['时域图谱位置']
                df_ssb['frequency'] = df_v['NRFreq的SSB的频点']
                df_ssb['subcarrierSpacing'] = df_v['子载波间隔'].apply(lambda x: '[' + x + ']')
                df_ssb.drop_duplicates(inplace=True)
                df_ssb = df_ssb[df_ssb['sourceGNBId'] != df_ssb['targetGNBId']]
                sheet = wb['SsbMeasInfo']
                r = 6
                for va in [j.tolist() for j in df_ssb.values]:
                    c = 1
                    for vv in va:
                        sheet.cell(r, c).value = vv
                        c += 1
                    r += 1

                df_nei = pd.DataFrame(
                    columns=["MODIND", "ManagedElementType", "sourcePlmn", "sourceGNBId", "sourceCellId",
                             'targetPlmn', "targetGNBId", 'targetCellId', 'userLabel'])

                df_nei['sourcePlmn'] = df_v['S_PLMN列表'].apply(lambda x: x.split(';')[0])
                df_nei['sourceGNBId'] = df_v['GNBID']
                df_nei['MODIND'] = 'A'
                df_nei['ManagedElementType'] = 'ITBBU'
                df_nei['sourceCellId'] = df_v['小区标识']
                df_nei['targetPlmn'] = df_v['PLMN列表'].apply(lambda x: x.split(';')[0])
                df_nei['targetGNBId'] = df_v['N_enb']
                df_nei['targetCellId'] = df_v['N_cell']
                df_nei['userLabel'] = df_v['用户标识']

                df_nei.drop_duplicates(inplace=True)
                sheet = wb['NRCellRelation']
                r = 6
                for va in [j.tolist() for j in df_nei.values]:
                    c = 1
                    for vv in va:
                        sheet.cell(r, c).value = vv
                        c += 1
                    r += 1
                wb.save(os.path.join(output, 'RANCM-addcellrelation_nr2nr.xlsx'))
        else:
            print(f"表 {i} 没有数据。")


if __name__ == '__main__':

    path = r'E:\Desktop\测评\现场模版样式20240421\1-4&G网管脚本的小区必须信息.xlsx'
    import_cell(path)
    import_nei('add.xlsx')

    do_add_nei('')
