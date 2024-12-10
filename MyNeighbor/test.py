import cv2
import numpy as np
img = cv2.imread('1.png')
new = np.clip(1.4057577998008846*img-0, 0, 255).astype(np.uint8)
cv2.imwrite('removed.png', new)

from openpyxl import load_workbook

# 加载已存在的Excel工作簿
workbook = load_workbook('Excel_addcellrelation_lte.xlsm',keep_vba=True)
print(workbook.sheetnames)

# 选择要写入数据的工作表
sheet = workbook['ExternalEUtranCellFDD']  # 或者使用workbook['工作表名']

# 向单元格写入数据
# sheet['A1'] = 'Hello, World!'

# 向下一个空白行写入多个数据
rows = [(i + 1, i + 2, i + 3) for i in range(10)]
for row in rows:
    sheet.append(row)

# 保存修改后的工作簿
workbook.save('Excel_addcellrelation_lte.xlsm')