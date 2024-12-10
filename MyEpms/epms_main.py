import json
import os
import datetime
import time
import sys
import traceback
from urllib.parse import quote, unquote
import openpyxl
import pandas as pd
import requests
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (QDialog, QAbstractItemView, QFileDialog, QApplication, QMessageBox, QLineEdit)
from openpyxl.utils import column_index_from_string, get_column_letter

from MyEpms.epms_login import epms_token
from MyEpms.ui_epms import Ui_Dialog_emps
from openpyxl.styles import Alignment, Font, Side, PatternFill, GradientFill, Border, numbers


class epms:
    def __init__(self, Account, Token):
        self.UCSSSOAccount = Account
        self.UCSSSOToken = Token
        self.userUrl = f'https://iepms.zte.com.cn/zte-crm-iepms-basebff/zte-crm-iepms-bcenter/api/users/{Account}'
        self.projectUrl = 'https://iepms.zte.com.cn/zte-crm-iepms-basebff/zte-crm-iepms-bcenter/projectinfoext/getpage'
        self.duModelUrl = 'https://iepms.zte.com.cn/zte-crm-iepms-basebff/zte-crm-iepms-schedule/duModel'
        self.dataUrl = 'https://iepms.zte.com.cn/zte-crm-iepms-basebff/zte-crm-iepms-schedule/schedule/getScheduleData'
        self.deviceUrl = 'https://iepms.zte.com.cn/zte-crm-iepms-basebff/zte-crm-iepms-schedule/schedule/exportDevice'
        self.titleUrl = 'https://iepms.zte.com.cn/zte-crm-iepms-basebff/zte-crm-iepms-schedule/schedule/getScheduleTitle?duModelId='
        self.header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
            'X-Emp-No': f'{Account}',
            'X-Auth-Value': f'{Token}'
        }
        self.title = None
        self.Itp = None
        if not os.path.exists('./download'):
            os.mkdir(r'./download')
        if not os.path.exists('./device'):
            os.mkdir(r'./device')
        self.save = []
        self.save_device = []

    def get_project(self):
        """
        获取项目ID
        :return:
        """
        data = {'pageNo': 1, 'pageSize': 200, 'projectDataType': "Y"}
        response = requests.post(url=self.projectUrl, headers=self.header, json=data)
        content = json.loads(response.content)
        project_dic = {}
        project_lst = []
        if content['code']['code'] == '0000':
            total = content['bo']['total']
            for i in range(int(total / 200 + 1)):
                data = {'pageNo': i + 1, 'pageSize': 200, 'projectDataType': "Y"}
                response = requests.post(url=self.projectUrl, headers=self.header, json=data)
                content = json.loads(response.content)
                if content['bo']['rows']:
                    dic_code = {i['projCode']: i['projId'] for i in content['bo']['rows']}
                    dic_Name = {i['projName']: i['projId'] for i in content['bo']['rows']}
                    project_lst += ([i['projName'] + ':' + i['projCode'] for i in content['bo']['rows']])
                    project_dic.update(dic_code)
                    project_dic.update(dic_Name)
                    project_dic.update({i['projId']: f"{i['projName']}-{i['projCode']}" for i in content['bo']['rows']})
                # print(project_dic)
        else:
            print(content['code'])
        return project_dic, project_lst

    def get_du(self, projName):
        """
        根据项目名称获取子项目ID
        :param projName:
        :return:
        """
        project_dic, project_lst = self.get_project()
        if project_dic and project_dic.get(projName):
            projId = project_dic.get(projName)
            projNameCode = project_dic.get(projId)
            self.Itp = {'X-Itp-Value': f'timeZone=8;projId={projId}'}
            self.header.update(self.Itp)
            response = requests.get(url=self.duModelUrl, headers=self.header)
            content = json.loads(response.content)
            duModelId = {}
            if content['code']['code'] == '0000':
                if content['bo']:
                    dic = {i['duModelName']: i['duModelId'] for i in content['bo'][0]['duModelVOList']}
                    duModelId.update(dic)
                else:
                    print(projName)
            return projId, duModelId, projNameCode
        else:
            print('请输入正确的项目名称')
            return None, None

    def get_title(self, duModelId):
        self.titleUrl += duModelId
        response = requests.get(url=self.titleUrl, headers=self.header, json={'duModelId': duModelId}, timeout=5)
        content = json.loads(response.content)
        title3 = {}
        if content['code']['code'] == '0000':
            title3.update({i['fieldPath']: i['titleTwo'] + '_' + i['titleThree'] for i in content['bo']})

        self.title = title3
        return [title3]

    def get_data(self, projNameCode, projId, duModelId_name, duModelId):
        """
        获取数据
        :param projNameCode: 项目编号
        :param projId: 项目ID
        :param duModelId_name:子项目名称
        :param duModelId: 子项目ID
        :return:
        """
        # projId, duModelId, projNameCode = self.get_du(projName)

        # self.title = self.get_title(list(duModelId.values())[0])  # 获取3级表头
        # print(self.title)
        lst = [self.title]
        print(projNameCode, projId, duModelId_name, duModelId)
        data = {'duModelId': duModelId, 'duStatus': '', 'pageNum': 1, 'pageSize': 500}
        response = requests.post(url=self.dataUrl, headers=self.header, json=data)
        content = json.loads(response.content)
        if content['code']['code'] == '0000':
            total = content['bo']['total']
            for i in range(int(total / 500 + 1)):
                data = {'duModelId': duModelId, 'duStatus': '', 'pageNum': i + 1, 'pageSize': 500}
                response = requests.post(url=self.dataUrl, headers=self.header, json=data)
                content = json.loads(response.content)
                if content['bo']['rows']:
                    for row in content['bo']['rows']:
                        dic = {j['fieldPath']: j['value'] for j in row['fieldVOList']}
                        lst.append(dic)
                # print(content)
            df = pd.DataFrame(lst)
            # df = df.T.set_index(0).T.reset_index(drop=True).rename_axis(columns=None)
            df.to_excel(f"./download/{projNameCode}-{duModelId_name}.xlsx", index=False)
            self.save.append(f"./download/{projNameCode}-{duModelId_name}.xlsx")

    def get_device(self, duModelId, projNameCode, duModelId_name):
        """
        :param duModelId: 子项目ID
        :param projNameCode: 项目编号
        :param duModelId_name: 子项目名称
        :return:
        """
        data = {'duModelId': duModelId, 'duStatus': "ENABLED", 'pageNum': 1, 'pageSize': 20, 'searchType': 'HIGH'}
        response = requests.post(url=self.deviceUrl, headers=self.header, json=data)
        time.sleep(5)
        content = json.loads(response.content)
        if content['code']['code'] == '0000':
            down_url = r'https://iepms.zte.com.cn/zte-crm-iepms-basebff/zte-crm-iepms-schedule/record?operationType=EXPORT&pageNo=1&pageSize=20&bizType=SCHEDULE'
            response = requests.get(url=down_url, headers=self.header)
            content = json.loads(response.content)
            if content['code']['code'] == '0000':
                total = content['bo']['total']
                if content['bo']['rows']:
                    while True:
                        if content['bo']['rows'][0]['fileSize'] == '0.00KB':
                            time.sleep(5)
                            response = requests.get(url=down_url, headers=self.header)
                            content = json.loads(response.content)
                            continue
                        else:
                            fileName = quote(content['bo']['rows'][0]['fileName'])  # 中文加密
                            fileId = content['bo']['rows'][0]['fileId']
                            excel_url = f'https://iepms.zte.com.cn/zte-crm-iepms-basebff/zte-crm-iepms-schedule/record/download?docId={fileId}&fileName={fileName}'
                            response = requests.get(url=excel_url, headers=self.header)
                            self.save_device.append(f"./device/{projNameCode}-{duModelId_name}.xlsx")
                            f = open(f"./device/{projNameCode}-{duModelId_name}.xlsx", "wb")
                            f.write(response.content)
                            f.close()
                            break


# 下载数据
def run_epms(account, token, projName):
    """
    :param account: 账号
    :param token: 密码
    :param projName: 项目名称
    :return:
    """
    myApp = epms(account, token)
    projId, duModelId, projNameCode = myApp.get_du(projName)
    Itp = myApp.Itp
    save_lst = []
    save_lst_device = []
    for k, v in duModelId.items():
        myApp = epms(account, token)  # 重新加载app，解决多次请求，响应消息不完成问题
        myApp.header.update(Itp)
        myApp.get_title(v)  # 获取表头
        # 只下载有单验的项目
        if '单验通过_实际结束' in myApp.title.values() and 'RRU安装完成_活动状态' in myApp.title.values():
            myApp.get_data(projNameCode, projId, k, v)  # 获取数据
            myApp.get_device(v, projNameCode, k)  # 获取数据
            save_lst += myApp.save
            save_lst_device += myApp.save_device
    return save_lst, save_lst_device


def count_define(df, key1):
    result = df.pivot_table(index=['站点基础信息_区域'], columns=[key1], values=['站点基础信息_客户站点编码'], aggfunc='count')
    cols_name = [f'{key1}_{i}' for i in result.columns.levels[1].to_list()]
    result.columns = cols_name
    result = result.reset_index()
    return result


def sum_define(df, key1):
    result = df.pivot_table(index=['区域名称'], columns=[key1], values=['设备数量'], aggfunc='sum')
    cols_name = [f'{i}' for i in result.columns.levels[1].to_list()]
    result.columns = cols_name
    result = result.reset_index()
    return result


class MyEpmsAnalyse(Ui_Dialog_emps, QDialog):
    def __init__(self):
        super().__init__()
        self.setAttribute(Qt.WidgetAttribute.WA_QuitOnClose, False)
        self.setupUi(self)
        self.show()
        self.username = ''
        # self.username = 'kongweibiao@shunscom.com'
        self.password = ''
        # self.password = 'Kwb231027'
        self.listWidget_project.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.lineEdit_password.setEchoMode(QLineEdit.EchoMode.Password)
        self.account = None
        self.token = None
        self.output_dir = ''
        self.pushButton_login.clicked.connect(self.login)
        self.pushButton_remove_proj.clicked.connect(self.remove_project)
        self.pushButton_do.clicked.connect(self.download)
        self.pushButton_get_proName.clicked.connect(self.get_project)

    def login(self):
        self.username = self.lineEdit_username.text()
        self.password = self.lineEdit_password.text()
        if not self.username:
            QMessageBox.warning(self, "信息提示", "请输入用户名")
            return
        if not self.password:
            QMessageBox.warning(self, "信息提示", "请输入密码")
            return
        try:
            app = epms_token(self.username, self.password)
            app.start()
        except:
            print(traceback.format_exc())
        self.account = app.account
        self.token = app.token
        if self.account:
            self.textBrowser_log.append(f'{datetime.datetime.now()}:{self.username}登录成功')
            QApplication.processEvents()
        else:
            self.textBrowser_log.append(f'{datetime.datetime.now()}:{self.username}登录失败')
            QApplication.processEvents()

    def get_project(self):
        if self.account:
            try:
                myApp = epms(self.account, self.token)
                dic, lst = myApp.get_project()
                self.listWidget_project.addItems(lst)
            except:
                print(traceback.format_exc())

    def remove_project(self):
        selected_items = self.listWidget_project.selectedItems()
        for item in selected_items:
            row = self.listWidget_project.row(item)
            self.listWidget_project.takeItem(row)

    def analyse(self, proj_paths):
        bold_11_red_font = Font(name='微软雅黑', size=11, italic=False, color='FF0000', bold=True)
        bold_11_black_font = Font(name='微软雅黑', size=11, italic=False, color='000000', bold=True)
        bold_14_black_font = Font(name='微软雅黑', size=14, italic=False, color='000000', bold=True)
        NonBold_11_black_font = Font(name='微软雅黑', size=11, italic=False, color='000000', bold=False)
        GradientFill1 = GradientFill(type='linear', degree=0, stop=('0070c0', 'FFFFFF', '0070c0'))  # 渐变颜色，16进制rgb
        fill1 = PatternFill(fill_type='solid', fgColor="C0C0C0", bgColor="C0C0C0")
        fill2 = PatternFill(fill_type='solid', fgColor="B0E0E6", bgColor="B0E0E6")
        fill3 = PatternFill(fill_type='solid', fgColor="92D050", bgColor="92D050")
        fill4 = PatternFill(fill_type='solid', fgColor="CCC0DA", bgColor="CCC0DA")
        fill5 = PatternFill(fill_type='solid', fgColor="FABF8F", bgColor="FABF8F")
        side = Side(style="thin", color="000000")
        all_Border = Border(left=side, right=side, top=side, bottom=side)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        row_no = 1
        cell = sheet.cell(row=row_no, column=1)
        cell.value = f"南京顺盛江苏区域重点工程进展情况"
        sheet.merge_cells(range_string=f'A{row_no}:U{row_no}')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = bold_14_black_font
        cell.fill = fill1
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = all_Border

        xlsx_dic = {}
        for xlsx in proj_paths:
            if xlsx.rsplit('_', 1)[0].replace('(RRU)', '') not in xlsx_dic.keys():
                xlsx_dic[xlsx.rsplit('_', 1)[0].replace('(RRU)', '')] = pd.read_excel(xlsx, header=1)
            else:
                xlsx_dic[xlsx.rsplit('_', 1)[0].replace('(RRU)', '')] = \
                    pd.concat([xlsx_dic[xlsx.rsplit('_', 1)[0].replace('(RRU)', '')], pd.read_excel(xlsx, header=1)],
                              join='outer', axis=0)

        for k, data in xlsx_dic.items():
            if data.shape[0] > 0:
                data.rename(columns={'RRU安装完成_实际结束': '安装完成_实际结束'}, inplace=True)
                # res = data.pivot_table(index=['站点基础信息_区域'], values=['安装完成_实际结束', '开通完成_实际结束', '单验通过_实际结束'],
                #                        aggfunc='count').fillna(0).reset_index()
                res0 = data.pivot_table(index=['站点基础信息_区域'], values=['站点基础信息_客户站点编码'], aggfunc='count').reset_index()
                res0.rename(columns={'站点基础信息_客户站点编码': '总数'}, inplace=True)
                res1 = count_define(data, '站点基础信息_交付单元状态')
                if '站点基础信息_交付单元状态_Enabled' not in res1.columns:
                    res1['站点基础信息_交付单元状态_Enabled'] = 0
                if '站点基础信息_交付单元状态_Disabled' not in res1.columns:
                    res1['站点基础信息_交付单元状态_Disabled'] = 0
                res2 = count_define(data, 'RRU安装完成_活动状态')
                if 'RRU安装完成_活动状态_Completed' not in res2.columns:
                    res2['RRU安装完成_活动状态_Completed'] = 0
                if 'RRU安装完成_活动状态_NoStart' not in res2.columns:
                    res2['RRU安装完成_活动状态_NoStart'] = 0
                if 'RRU安装完成_活动状态_OnGoing' not in res2.columns:
                    res2['RRU安装完成_活动状态_OnGoing'] = 0
                if 'RRU安装完成_活动状态_Abolished' not in res2.columns:
                    res2['RRU安装完成_活动状态_Abolished'] = 0

                res3 = count_define(data, '开通完成_活动状态')
                if '开通完成_活动状态_Completed' not in res3.columns:
                    res3['开通完成_活动状态_Completed'] = 0
                if '开通完成_活动状态_NoStart' not in res3.columns:
                    res3['开通完成_活动状态_NoStart'] = 0
                if '开通完成_活动状态_OnGoing' not in res3.columns:
                    res3['开通完成_活动状态_OnGoing'] = 0
                if '开通完成_活动状态_Abolished' not in res3.columns:
                    res3['开通完成_活动状态_Abolished'] = 0

                res4 = count_define(data, '单验通过_活动状态')
                if '单验通过_活动状态_Completed' not in res4.columns:
                    res4['单验通过_活动状态_Completed'] = 0
                if '单验通过_活动状态_NoStart' not in res4.columns:
                    res4['单验通过_活动状态_NoStart'] = 0
                if '单验通过_活动状态_OnGoing' not in res4.columns:
                    res4['单验通过_活动状态_OnGoing'] = 0
                if '单验通过_活动状态_Abolished' not in res4.columns:
                    res4['单验通过_活动状态_Abolished'] = 0

                res = res0.merge(res1, on='站点基础信息_区域', how='outer')
                res = res.merge(res2, on='站点基础信息_区域', how='outer')
                res = res.merge(res3, on='站点基础信息_区域', how='outer')
                res = res.merge(res4, on='站点基础信息_区域', how='outer')
                res['RRU安装未完成数量'] = res['总数'] - res['RRU安装完成_活动状态_Completed']
                res['开通未完成数量'] = res['总数'] - res['开通完成_活动状态_Completed']
                res['单验未完成数量'] = res['总数'] - res['单验通过_活动状态_Completed']

                # res['阶段'] = res['阶段'] + '-' + xlsx.rsplit('-')[-1].split('.xlsx')[0]

                res['开通率EPMS'] = res.apply(lambda x: x['开通完成_活动状态_Completed'] / x['RRU安装完成_活动状态_Completed'] if x[
                                                                                                                   'RRU安装完成_活动状态_Completed'] != 0 else 0,
                                           axis=1)
                res['单验率EPMS'] = res.apply(lambda x: x['单验通过_活动状态_Completed'] / x['开通完成_活动状态_Completed'] if x[
                                                                                                                '开通完成_活动状态_Completed'] != 0 else 0,
                                           axis=1)
                res = res[
                    ['站点基础信息_区域', '总数', '站点基础信息_交付单元状态_Disabled', '站点基础信息_交付单元状态_Enabled', 'RRU安装完成_活动状态_Abolished',
                     'RRU安装完成_活动状态_Completed', 'RRU安装完成_活动状态_NoStart', 'RRU安装完成_活动状态_OnGoing', 'RRU安装未完成数量',
                     '开通完成_活动状态_Abolished', '开通完成_活动状态_Completed', '开通完成_活动状态_NoStart', '开通完成_活动状态_OnGoing', '开通未完成数量',
                     '单验通过_活动状态_Abolished', '单验通过_活动状态_Completed', '单验通过_活动状态_NoStart', '单验通过_活动状态_OnGoing', '单验未完成数量',
                     '开通率EPMS', '单验率EPMS']]

                row_no += 1
                cell = sheet.cell(row=row_no, column=1)
                cell.value = f"{k.rsplit('-', 2)[0].split('/')[-1]}({k.split('-')[-1]})"
                sheet.merge_cells(range_string=f'A{row_no}:U{row_no}')
                cell.font = bold_11_red_font
                cell.fill = GradientFill1
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = all_Border
                row_no += 1
                for col in range(res.shape[1]):
                    cell = sheet.cell(row=row_no, column=col + 1)
                    cell.value = list(res.columns)[col]
                    cell.font = bold_11_black_font
                    cell.fill = fill2
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = all_Border
                for index, row in res.iterrows():
                    value_lst = row.to_list()
                    row_no += 1
                    for col in range(res.shape[1]):
                        cell = sheet.cell(row=row_no, column=col + 1)
                        cell.value = value_lst[col]
                        cell.font = NonBold_11_black_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if col + 1 in [4, 6, 11, 16, 20, 21]:
                            cell.fill = fill3
                        if col + 1 in [3, 5, 10, 15]:
                            cell.fill = fill4
                        if col + 1 in [9, 14, 19]:
                            cell.fill = fill5
                        cell.border = all_Border

        """for xlsx in proj_paths:
            data = pd.read_excel(xlsx)
            data = data.T.set_index(0).T.reset_index(drop=True).rename_axis(columns=None)
            if data.shape[0] > 0:
                data.rename(columns={'RRU安装完成_实际结束': '安装完成_实际结束'}, inplace=True)
                res = data.pivot_table(index=['站点基础信息_区域'], values=['安装完成_实际结束', '开通完成_实际结束', '单验通过_实际结束'],
                                       aggfunc='count').fillna(0).reset_index()
                res.rename(
                    columns={'站点基础信息_区域': '地市', '安装完成_实际结束': '安装EPMS', '开通完成_实际结束': '开通EPMS', '单验通过_实际结束': '单验EPMS'},
                    inplace=True)
                print(xlsx, '\n', res)

                # res['阶段'] = res['阶段'] + '-' + xlsx.rsplit('-')[-1].split('.xlsx')[0]
                res['开通率EPMS'] = res.apply(lambda x: x['开通EPMS'] / x['安装EPMS'] if x['安装EPMS'] != 0 else 0, axis=1)
                res['单验率EPMS'] = res.apply(lambda x: x['单验EPMS'] / x['开通EPMS'] if x['开通EPMS'] != 0 else 0, axis=1)
                row_no += 1
                cell = sheet.cell(row=row_no, column=1)
                cell.value = f"{xlsx.rsplit('-', 2)[0].split('.xlsx')[0].rsplit(r'/', 1)[-1]}({xlsx.split('-')[-1].split('.xlsx')[0]})"
                sheet.merge_cells(range_string=f'A{row_no}:F{row_no}')
                cell.font = bold_11_red_font
                cell.fill = GradientFill1
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = all_Border
                row_no += 1
                for col in range(res.shape[1]):
                    cell = sheet.cell(row=row_no, column=col + 1)
                    cell.value = list(res.columns)[col]
                    cell.font = bold_11_black_font
                    cell.fill = fill2
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = all_Border
                for index, row in res.iterrows():
                    value_lst = row.to_list()
                    row_no += 1
                    for col in range(res.shape[1]):
                        cell = sheet.cell(row=row_no, column=col + 1)
                        cell.value = value_lst[col]
                        cell.font = NonBold_11_black_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = all_Border"""

        col1 = column_index_from_string('T')
        col2 = column_index_from_string('U')

        for i in range(sheet.max_row):
            cell1 = sheet.cell(row=1 + i, column=col1)
            cell2 = sheet.cell(row=1 + i, column=col2)
            cell1.number_format = numbers.FORMAT_PERCENTAGE
            cell2.number_format = numbers.FORMAT_PERCENTAGE
        for n in range(1, 18):
            sheet.column_dimensions[get_column_letter(n)].width = 11
        save_name = os.path.join(self.output_dir, f'epms统计结果-{time.strftime("%Y%m%d")}.xlsx')
        workbook.save(save_name)
        self.textBrowser_log.append(f'分析完成，保存路径为{save_name}')
        QApplication.processEvents()
        return sheet

    def analyse_device(self, proj_paths_device):
        side = Side(style="thin", color="000000")
        all_Border = Border(left=side, right=side, top=side, bottom=side)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # row_no = 1
        cell = sheet.cell(row=1, column=1)
        cell.value = f"项目"
        sheet.merge_cells(range_string='A1:A2')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = all_Border
        cell = sheet.cell(row=2, column=1)
        cell.border = all_Border
        cell = sheet.cell(row=1, column=2)
        cell.value = f"地市"
        sheet.merge_cells(range_string='B1:B2')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = all_Border
        cell = sheet.cell(row=1, column=3)
        cell.value = f"epms安装-设备级"
        sheet.merge_cells(range_string='C1:D1')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = all_Border
        cell = sheet.cell(row=2, column=3)
        cell.value = f"RRU/AAU"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = all_Border
        cell = sheet.cell(row=2, column=4)
        cell.value = f"PRRU"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = all_Border
        cell = sheet.cell(row=1, column=5)
        cell.value = f"epms优化-设备级"
        sheet.merge_cells(range_string='E1:F1')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = all_Border
        cell = sheet.cell(row=2, column=5)
        cell.value = f"RRU/AAU"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = all_Border
        cell = sheet.cell(row=2, column=6)
        cell.value = f"PRRU"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = all_Border
        row_no = 2

        xlsx_dic = {}
        for xlsx in proj_paths_device:
            if xlsx.rsplit('-', 1)[0] not in xlsx_dic.keys():
                xlsx_dic[xlsx.rsplit('-', 1)[0]] = pd.read_excel(xlsx)
            else:
                xlsx_dic[xlsx.rsplit('-', 1)[0]] = \
                    pd.concat([xlsx_dic[xlsx.rsplit('-', 1)[0]], pd.read_excel(xlsx)], join='outer', axis=0)

        for k, data in xlsx_dic.items():
            if data.shape[0] > 0:
                data_RRU = data[(data['活动状态'] == '已完成') & (data['活动名称'] == 'RRU安装完成')]
                RRU = data_RRU.pivot_table(index=['客户站点编码', '设备类型', '设备型号'], values=['设备数量'], aggfunc='sum').fillna(0)
                RRU = RRU.reset_index()
                data_dany = data[(data['活动状态'] == '已完成') & (data['活动名称'] == '单验通过')]
                data_dany = data_dany.merge(RRU, on='客户站点编码', how='left', suffixes=('_x', '')).fillna(0)

                res_rru = sum_define(data_RRU, '设备类型')
                if 'pRRU' not in res_rru.columns:
                    res_rru['pRRU'] = 0
                if 'AAU/RRU' not in res_rru.columns:
                    res_rru['AAU/RRU'] = 0

                res_dany = sum_define(data_dany, '设备类型')
                if 'pRRU' not in res_dany.columns:
                    res_dany['pRRU'] = 0
                if 'AAU/RRU' not in res_dany.columns:
                    res_dany['AAU/RRU'] = 0

                res = res_rru.merge(res_dany, on='区域名称', how='left', suffixes=('', '_y')).fillna(0)
                res['项目'] = k.split('-')[0].split('/')[-1]

                for index, row in res.iterrows():
                    row_no += 1
                    cell = sheet.cell(row=row_no, column=1)
                    cell.value = row['项目']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = all_Border
                    cell = sheet.cell(row=row_no, column=2)
                    cell.value = row['区域名称']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = all_Border
                    cell = sheet.cell(row=row_no, column=3)
                    cell.value = row['AAU/RRU']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = all_Border
                    cell = sheet.cell(row=row_no, column=4)
                    cell.value = row['pRRU']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = all_Border
                    cell = sheet.cell(row=row_no, column=5)
                    cell.value = row['AAU/RRU_y']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = all_Border
                    cell = sheet.cell(row=row_no, column=6)
                    cell.value = row['pRRU_y']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = all_Border

        save_name = os.path.join(self.output_dir, f'epms-device-统计结果-{time.strftime("%Y%m%d")}.xlsx')
        workbook.save(save_name)
        self.textBrowser_log.append(f'分析完成，保存路径为{save_name}')
        QApplication.processEvents()
        return sheet

    def download(self):
        if self.account:
            self.output_dir = QFileDialog.getExistingDirectory(self, "请选择保存目录", os.getcwd())
            proj_paths = []
            proj_paths_device = []
            for idx in range(self.listWidget_project.count()):
                item = self.listWidget_project.item(idx)
                # proj_paths.append(item.text())
                self.textBrowser_log.append(f'{datetime.datetime.now()}:正在下载《{item.text().split(":")[0]}》项目文件')
                QApplication.processEvents()
                try:
                    sve, save_device = run_epms(self.account, self.token, item.text().split(':')[0])
                    proj_paths += sve
                    proj_paths_device += save_device
                    # print(proj_paths)
                except:
                    self.textBrowser_log.append(f'{datetime.datetime.now()}:下载失败《{item.text().split(":")[0]}》项目文件')
                    QApplication.processEvents()
                    print(traceback.format_exc())
            if proj_paths:
                self.textBrowser_log.append(f'{datetime.datetime.now()}:数据分析中...')
                QApplication.processEvents()
                try:
                    self.analyse(proj_paths)
                    self.analyse_device(proj_paths_device)
                except:
                    print(traceback.format_exc())
                    self.textBrowser_log.append(f'分析完成，保存路径为{traceback.format_exc()}')
                    QApplication.processEvents()

        else:
            self.textBrowser_log.append(f'{datetime.datetime.now()}:{self.username}登录失败')
            QApplication.processEvents()


# 按装订区域中的绿色按钮以运行脚本。
if __name__ == '__main__':
    # username = 'kongweibiao@shunscom.com'
    # password = 'Kwb231027'
    # app = epms_token(username, password)
    # app.start()
    # if app.account:
    #     print(app.account, app.token)
    #     run_epms(app.account, app.token,'P202201163972_D005')
    # else:
    #     print('滑动验证失败')
    app1 = QApplication(sys.argv)

    myEpmsAnalyse = MyEpmsAnalyse()
    sys.exit(app1.exec())
